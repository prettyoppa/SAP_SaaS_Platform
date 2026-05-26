"""선불 충전·마켓플레이스 정산 전략 설명을 Excel(.xlsx)로보냅니다."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "prepaid_marketplace_strategy.xlsx"


def _sheet(wb, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title=title[:31])
    bold = Font(bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(48, max(12, len(headers[c - 1]) + 2))
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    _sheet(
        wb,
        "00_요약",
        ["항목", "권장/결론"],
        [
            ["구독 → 선불 충전", "찬성 — 간헐적·프로젝트성 사용 패턴에 적합"],
            ["플랜별 기능 잠금 제거", "찬성 — 잔액 + 역할(member/consultant/admin)만 유지"],
            ["마켓플레이스 정산", "찬성 — RequestOffer/match 기반으로 자연스러운 확장"],
            ["한 번에 전부 구현", "비권장"],
            ["PortOne/Stripe Connect 즉시", "비권장 — 수동 정산 검증 후 Phase 4"],
            ["추천 순서", "Phase 0 → Phase 1(지갑) → Phase 2(에스크로·수동) → Phase 3~4"],
            ["핵심 설계 원칙", "AI 플랫폼 이용료(A)와 컨설턴트 프로젝트 대금(B) 원장 분리"],
        ],
    )

    _sheet(
        wb,
        "01_비즈니스_의견",
        ["주제", "내용"],
        [
            ["서비스 특성", "매월 반복 SaaS보다 요청 발생 시 집중 사용(인터뷰·제안·FS·납품)"],
            ["구독제 부담", "안 쓰는 달에도 비용·플랜별 기능 잠금으로 마찰"],
            ["선불 충전 모델", "잔액 있으면 사용 / 없으면 충전 — mental model 단순"],
            ["주의: 예측 가능성", "LLM 토큰 변동 → UI에 예상 차감액 표시 필요(ai_usage_events 활용)"],
            ["주의: 충전·환불", "최소 충전액·미사용 잔액 환불 — 약관 명시"],
            ["주의: 혼동 방지", "플랫폼 AI 이용료 vs 컨설턴트 인건비 UI/문구 분리"],
            ["역할 유지", "member / consultant / admin"],
            ["역할 제거(과금)", "Junior/Senior/Experience 등 플랜 등급·entitlement"],
        ],
    )

    _sheet(
        wb,
        "02_현재코드_재활용",
        ["영역", "현재 구현", "전환 시 활용"],
        [
            ["계좌이체", "PaymentClaim, /account/billing, /admin/payment-claims", "충전 신청·승인 → 잔액 적립"],
            ["AI 비용 추정", "AiUsageEvent, aggregate_usage_for_user()", "차감 단가·리포트 기초"],
            ["월간 metric", "SubscriptionUsageMonthly", "한도 대신 표시/리포트(선택)"],
            ["컨설턴트 매칭", "RequestOffer (offered/matched/withdrawn)", "마켓 거래·에스크로 업무 단위"],
            ["RFP 유료 납품", "paid_engagement_status + Stripe 일회성", "프로젝트 대금 흐름과 통합 여부 결정 필요"],
        ],
    )

    _sheet(
        wb,
        "03_현재코드_미구현",
        ["항목", "설명"],
        [
            ["지갑 원장", "잔액·충전·차감·환불·통화"],
            ["기능 게이트", "entitlement → 잔액 게이트 일괄 교체"],
            ["카드 충전", "구독이 아닌 top-up PaymentIntent"],
            ["프로젝트 금액", "수수료·컨설턴트 지급액 모델"],
            ["에스크로 상태", "결제 보류 → 납품/확인 → 지급"],
            ["정산 프로필", "컨설턴트 계좌·Stripe Connect ID·세금정보"],
        ],
    )

    _sheet(
        wb,
        "04_결제축_분리",
        ["축", "명칭", "설명", "비고"],
        [
            ["A", "플랫폼 지갑", "AI·자동화 사용료 — 토큰/건당 잔액 차감", "회원 충전"],
            ["B", "프로젝트 에스크로", "수요자 ↔ 컨설턴트 인건비 — 매칭·납품 연동", "마켓플레이스"],
            ["C", "플랫폼 수수료", "B에서 % 또는 별도", "정책 결정"],
            ["", "혼합 금지", "한 지갑에 A+B 합치면 환불·분쟁·회계 어려움", "UI 혼동"],
        ],
    )

    _sheet(
        wb,
        "05_기존_결제_혼선",
        ["갈래", "구현", "전환 시"],
        [
            ["1", "구독/계좌이체 — subscription_*, payment_claims", "Phase 1에서 지갑으로 대체"],
            ["2", "RFP Stripe 일회 — paid_engagement_status, payments_router", "프로젝트 에스크로(B)와 통합 또는 병행 결정"],
        ],
    )

    _sheet(
        wb,
        "06_PortOne_검토",
        ["제안 내용", "이 웹앱 적용", "비고"],
        [
            ["에스크로+분할 정산 프레임", "적합", ""],
            ["1단계 수동 → 2단계 API", "적합 — 거래량 적을 때"],
            ["KR PortOne + 해외 Stripe Connect", "MVP는 KRW·국내·수동 정산 우선", "해외는 수요 확인 후"],
            ["bridge/ 폴더", "app/payments/ 또는 payment_providers 모듈", "현재 레포 구조"],
            ["즉시 자동 분할", "비권장", "납품·분쟁 상태机 먼저"],
            ["구독 제거와 Connect 동시", "비권장", "Phase 1 지갑 먼저"],
            ["법무", "선불충전·에스크로·원천징수 — 약관·법무 1회 검토", ""],
        ],
    )

    _sheet(
        wb,
        "07_Phase0_정책",
        ["#", "결정 항목", "선택지/메모"],
        [
            ["1", "통화", "MVP KRW만 vs KRW+USD 지갑 분리"],
            ["2", "과금 단위", "ai_usage micro-USD→KRW 차감 vs 건당 정액"],
            ["3", "잔액 부족", "하드 차단 vs 일부 무료"],
            ["4", "프로젝트 가격", "제안 수락 시 고정 vs 협상 후 확정"],
            ["5", "지급 트리거", "matched / 납품완료 / 요청자확인 / N일 후 자동"],
            ["6", "수수료", "플랫폼 % + PG 수수료 부담 주체"],
            ["산출물", "ER 초안, 상태 다이어그램, 구독 폐기 목록", ""],
        ],
    )

    _sheet(
        wb,
        "08_Phase1_지갑",
        ["구분", "내용"],
        [
            ["목표", "플랜·월 한도 제거 → 충전 잔액으로 AI/플랫폼 사용"],
            ["테이블 예", "wallet_account — user_id, currency, balance_minor"],
            ["테이블 예", "wallet_ledger — type, amount, ref, balance_after, idempotency"],
            ["테이블 예", "topup_order — bank|card, status, pg_payment_id"],
            ["PaymentClaim", "topup 승인 시 ledger credit"],
            ["AI 사용", "log_ai_usage_event 후 debit"],
            ["게이트", "check_entitlement → assert_wallet_balance"],
            ["회원 UI", "/account/wallet (billing 대체)"],
            ["관리자 UI", "/admin/topups (payment-claims 개편)"],
            ["카드", "Stripe PaymentIntent 일회 충전"],
            ["계좌", "기존 수동 확인 유지"],
            ["구독 UI", "/subscription-plans → 단가·충전 안내로 변경"],
            ["마이그레이션", "기존 구독자 일회 크레딧(운영 정책)"],
        ],
    )

    _sheet(
        wb,
        "09_Phase2_마켓",
        ["구분", "내용"],
        [
            ["목표", "수요자 결제 → 완료 후 컨설턴트 수동 이체"],
            ["상태 예", "offered → matched → escrow_funded → in_delivery → released / disputed"],
            ["테이블 예", "market_order — gross, platform_fee, seller_net, status"],
            ["테이블 예", "payout_record — pending|paid, admin_note"],
            ["테이블 예", "consultant_payout_profile — 계좌·세금(암호화)"],
            ["결제 MVP-A", "지갑 hold → 컨설턴트 출금 가능 잔액 credit (개발량 적음)"],
            ["결제 MVP-B", "프로젝트만 별도 PG (회계 분리 명확)"],
            ["관리자", "/admin/settlements — pending 지급 완료 체크"],
            ["자동화", "없음 — 월말 계좌이체"],
        ],
    )

    _sheet(
        wb,
        "10_Phase3_4",
        ["Phase", "내용"],
        [
            ["Phase 3", "잔액 부족 전 예상 비용 UI"],
            ["Phase 3", "프로젝트 비용 상한(hold cap)"],
            ["Phase 3", "분쟁·부분환불·취소 ↔ offer lifecycle"],
            ["Phase 3", "컨설턴트 출금 요청(최소금액·수수료)"],
            ["Phase 4", "PortOne Partners — portone_partner_id"],
            ["Phase 4", "Stripe Connect Express — stripe_account_id"],
            ["Phase 4", "모듈: payment_providers/portone.py, stripe_connect.py"],
            ["전제", "Phase 2 market_order/payout 상태机 안정"],
        ],
    )

    _sheet(
        wb,
        "11_구독제거_순서",
        ["순서", "작업"],
        [
            ["1", "wallet_* + ledger + topup 신규"],
            ["2", "subscription_quota → wallet_debit / can_afford"],
            ["3", "UI: subscription-plans, billing, payment-claims → wallet"],
            ["4", "SubscriptionPlan → 단가표·수수료율 설정으로 재활용 가능"],
            ["5", "paid_engagement + Stripe ↔ 에스크로 통합 여부 결정"],
            ["6", "문서·i18n·관리자 메뉴 정리"],
        ],
    )

    _sheet(
        wb,
        "12_관리자_URL_현재",
        ["기능", "URL", "비고"],
        [
            ["관리자 홈", "/admin", "대시보드 카드"],
            ["입금 신청(계좌이체)", "/admin/payment-claims", "구독 활성화용 — 지갑 전환 대상"],
            ["회원 목록", "/admin/users", "가로 스크롤 → 구독 버튼"],
            ["회원별 구독·한도·AI비용", "/admin/users/{id}/subscription", "통합 모니터링 화면 없음"],
            ["구독 플랜 설정", "/admin/subscription-plans", "계좌 안내 포함"],
            ["회원 충전(회원)", "/account/billing", ""],
            ["전체 회원 사용량 대시", "(미구현)", "회원별로만 조회 가능"],
        ],
    )

    _sheet(
        wb,
        "13_Phase0_확정질문",
        ["#", "질문", "영향"],
        [
            ["1", "프로젝트비도 같은 충전 잔액?", "원장 A 단일 vs A+B 분리"],
            ["2", "MVP 통화 KRW만?", "PG·UI·환율"],
            ["3", "AI 과금: 원가+마진 vs 건당 정액?", "ai_usage 연동 vs 단순 요금"],
            ["4", "컨설턴트 지급 트리거?", "상태机·에스크로 해제 조건"],
        ],
    )

    _sheet(
        wb,
        "14_주요_파일_참고",
        ["파일/모듈", "역할"],
        [
            ["app/subscription_quota.py", "플랜 한도 — 제거/대체 대상"],
            ["app/subscription_catalog.py", "플랜·metric 정의"],
            ["app/payment_claim_service.py", "계좌이체 — topup으로 전환"],
            ["app/routers/billing_router.py", "회원 billing"],
            ["app/ai_usage_recorder.py", "AI 비용 이벤트"],
            ["app/models.py PaymentClaim, AiUsageEvent", "DB 모델"],
            ["app/paid_tier.py", "FS 접근 — 프로젝트 대금 여부로 재정의"],
            ["app/models.py RequestOffer", "마켓 매칭"],
            ["app/stripe_service.py", "Stripe — top-up·에스크로 재사용"],
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(str(OUT))


if __name__ == "__main__":
    main()
