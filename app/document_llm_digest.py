"""첨부·FS·제안서 파일 — 텍스트 추출 + 내장/단독 이미지 비전 요약."""

from __future__ import annotations

import io
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

from .codelib_bulk_upload import text_from_docx_bytes
from .gemini_vision_digest import vision_summarize_images
from .request_attachments import IMAGE_ATTACHMENT_EXTENSIONS

_MAX_DOC_EMBEDDED_IMAGES = 12
_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_OOXML_IMAGE_EXT = frozenset({".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"})


def _decode_text(raw: bytes, max_chars: int) -> str:
    t = ""
    for enc in ("utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"):
        try:
            t = raw.decode(enc)
            break
        except Exception:
            continue
    else:
        t = raw.decode("utf-8", errors="replace")
    t = t.replace("\x00", " ")
    if len(t) > max_chars:
        t = t[:max_chars] + "\n…(이하 잘림)…"
    return t


def _pdf_text_pypdf(raw: bytes, max_chars: int) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return "(PDF 파싱 라이브러리를 사용할 수 없습니다.)"
    try:
        r = PdfReader(io.BytesIO(raw))
        parts: list[str] = []
        remain = max_chars
        for page in r.pages:
            if remain <= 0:
                break
            try:
                txt = (page.extract_text() or "").strip()
            except Exception:
                txt = ""
            if not txt:
                continue
            if len(txt) > remain:
                txt = txt[:remain] + "\n…(페이지 잘림)…"
            parts.append(txt)
            remain -= len(txt)
        return "\n\n".join(parts).strip() if parts else ""
    except Exception as e:
        return f"(PDF 읽기 오류: {e})"


def _pdf_text_and_images(raw: bytes, max_chars: int) -> tuple[str, list[tuple[bytes, str]]]:
    images: list[tuple[bytes, str]] = []
    text = ""
    try:
        import fitz  # pymupdf
    except Exception:
        return _pdf_text_pypdf(raw, max_chars), images

    try:
        doc = fitz.open(stream=raw, filetype="pdf")
        text_parts: list[str] = []
        remain = max_chars
        for i, page in enumerate(doc):
            if remain > 0:
                try:
                    txt = (page.get_text() or "").strip()
                except Exception:
                    txt = ""
                if txt:
                    if len(txt) > remain:
                        txt = txt[:remain] + "\n…"
                    text_parts.append(txt)
                    remain -= len(txt)
            if len(images) >= _MAX_DOC_EMBEDDED_IMAGES:
                continue
            if not (page.get_text() or "").strip():
                try:
                    pix = page.get_pixmap(dpi=110, alpha=False)
                    png = pix.tobytes("png")
                    if png and len(png) >= 4096:
                        images.append((png, f"pdf-page-{i + 1}.png"))
                except Exception:
                    pass
            try:
                for img in page.get_images(full=True):
                    if len(images) >= _MAX_DOC_EMBEDDED_IMAGES:
                        break
                    xref = img[0]
                    base = doc.extract_image(xref)
                    if not base or not base.get("image"):
                        continue
                    ext = "." + (base.get("ext") or "png").lower()
                    if ext not in _OOXML_IMAGE_EXT:
                        continue
                    blob = base["image"]
                    if len(blob) >= 4096:
                        images.append((blob, f"pdf-img-p{i + 1}{ext}"))
            except Exception:
                pass
        doc.close()
        text = "\n\n".join(text_parts).strip()
    except Exception as e:
        text = _pdf_text_pypdf(raw, max_chars)
        if text.startswith("("):
            return text, images
    if not text:
        text = _pdf_text_pypdf(raw, max_chars)
    return text, images[:_MAX_DOC_EMBEDDED_IMAGES]


def _xlsx_outline(raw: bytes, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "(Excel 파싱 라이브러리를 사용할 수 없습니다.)"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        out: list[str] = []
        used = 0
        for sheet_name in wb.sheetnames:
            if used >= max_chars:
                break
            ws = wb[sheet_name]
            hdr = ["--- 시트 ---", sheet_name, "--- 샘플(최대 15행) ---"]
            lines: list[str] = list(hdr)
            for ri, row in enumerate(
                ws.iter_rows(min_row=1, max_row=15, max_col=40, values_only=True), start=1
            ):
                cells = []
                for c in row:
                    if c is None:
                        cells.append("")
                    else:
                        cells.append(str(c).replace("\n", " ").strip()[:120])
                lines.append(f"행{ri}: " + " | ".join(cells))
            block = "\n".join(lines)
            if used + len(block) > max_chars:
                block = block[: max_chars - used] + "\n…"
            out.append(block)
            used += len(block)
        wb.close()
        return "\n\n".join(out) if out else "(시트를 읽지 못했습니다.)"
    except Exception as e:
        return f"(Excel 읽기 오류: {e})"


def _ooxml_media_images(zf: zipfile.ZipFile, media_dir: str) -> list[tuple[bytes, str]]:
    prefix = media_dir.rstrip("/") + "/"
    out: list[tuple[bytes, str]] = []
    for name in zf.namelist():
        if not name.startswith(prefix):
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext not in _OOXML_IMAGE_EXT:
            continue
        try:
            blob = zf.read(name)
        except Exception:
            continue
        if len(blob) < 4096:
            continue
        out.append((blob, os.path.basename(name)))
        if len(out) >= _MAX_DOC_EMBEDDED_IMAGES:
            break
    return out


def _text_from_pptx_bytes(raw: bytes, max_chars: int) -> str:
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        return "(PPTX 형식이 아닙니다.)"
    slide_names = sorted(
        n for n in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", n, re.I)
    )
    parts: list[str] = []
    used = 0
    for sn in slide_names:
        if used >= max_chars:
            break
        try:
            root = ET.fromstring(zf.read(sn))
        except Exception:
            continue
        chunks: list[str] = []
        for el in root.iter():
            if el.tag.endswith("}t") and el.text and el.text.strip():
                chunks.append(el.text.strip())
        if chunks:
            block = " ".join(chunks)
            if used + len(block) > max_chars:
                block = block[: max_chars - used] + "…"
            parts.append(block)
            used += len(block)
    zf.close()
    return "\n\n".join(parts).strip() if parts else "(PPTX에서 추출한 텍스트가 없습니다.)"


def _docx_text_and_images(raw: bytes, max_chars: int) -> tuple[str, list[tuple[bytes, str]]]:
    text, err = text_from_docx_bytes(raw)
    images: list[tuple[bytes, str]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
        images = _ooxml_media_images(zf, "word/media")
        zf.close()
    except Exception:
        pass
    if err:
        return f"({err})", images
    if text and len(text) > max_chars:
        text = text[:max_chars] + "\n…(이하 잘림)…"
    return text or "", images


def _pptx_text_and_images(raw: bytes, max_chars: int) -> tuple[str, list[tuple[bytes, str]]]:
    images: list[tuple[bytes, str]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
        images = _ooxml_media_images(zf, "ppt/media")
        zf.close()
    except Exception:
        pass
    return _text_from_pptx_bytes(raw, max_chars), images


def _xlsx_text_and_images(raw: bytes, max_chars: int) -> tuple[str, list[tuple[bytes, str]]]:
    images: list[tuple[bytes, str]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
        images = _ooxml_media_images(zf, "xl/media")
        zf.close()
    except Exception:
        pass
    return _xlsx_outline(raw, max_chars), images


def document_file_digest(filename: str, raw: bytes, per_budget: int) -> str:
    """단일 파일 → LLM용 문자열(텍스트 + 선택적 비전 요약)."""
    fn = (filename or "file").strip()
    ext = os.path.splitext(fn)[1].lower()
    if not raw:
        return f"[{fn}] — 파일이 비어 있습니다."

    text_budget = max(800, int(per_budget * 0.55))
    vision_budget = max(400, per_budget - text_budget)
    text_part = ""
    images: list[tuple[bytes, str]] = []

    if ext in IMAGE_ATTACHMENT_EXTENSIONS:
        images = [(raw, fn)]
        text_part = ""
    elif ext in (".txt", ".md", ".csv", ".log", ".json", ".xml"):
        text_part = _decode_text(raw, text_budget)
    elif ext == ".pdf":
        text_part, images = _pdf_text_and_images(raw, text_budget)
    elif ext == ".docx":
        text_part, images = _docx_text_and_images(raw, text_budget)
    elif ext == ".pptx":
        text_part, images = _pptx_text_and_images(raw, text_budget)
    elif ext == ".xlsx":
        text_part, images = _xlsx_text_and_images(raw, text_budget)
    else:
        return f"[{fn}] 지원하지 않는 형식입니다."

    blocks: list[str] = [f"[{fn}]"]
    if text_part and text_part.strip():
        blocks.append(text_part.strip())
    elif not images:
        blocks.append("(추출한 텍스트가 없습니다.)")

    if images:
        vis = vision_summarize_images(
            images,
            max_chars=vision_budget,
            header=f"[{fn} — 문서·파일 내 이미지 시각 요약]\n",
        )
        if vis:
            blocks.append(vis)

    return "\n\n".join(blocks).strip()


def supplement_file_body_for_agents(filename: str, raw: bytes, *, max_chars: int = 48_000) -> tuple[str | None, str | None]:
    """
    FS·제안서 첨부 바이너리 → 에이전트용 본문 문자열.
    """
    if not raw:
        return None, "파일이 비어 있습니다."
    body = document_file_digest(filename, raw, max_chars)
    if not body or body.endswith("지원하지 않는 형식입니다."):
        return None, f"첨부를 해석하지 못했습니다: {filename}"
    if len(body) > max_chars:
        body = body[:max_chars] + "\n…(첨부 본문 상한)…"
    return body, None
