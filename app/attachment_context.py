"""첨부 파일에서 LLM 컨텍스트용 텍스트 추출(텍스트·PDF·Excel 개요)."""

from __future__ import annotations

import io
import os
from typing import Any

from . import r2_storage

# 확장자별 처리
_TEXT_EXT = frozenset({
    ".txt", ".csv", ".tsv", ".log", ".md", ".json", ".xml", ".yml", ".yaml", ".sql", ".ini", ".properties",
})
_EXCEL_EXT = frozenset({".xlsx", ".xlsm"})
_PDF_EXT = frozenset({".pdf"})


def _decode_text(raw: bytes, max_chars: int) -> str:
    t = ""
    for enc in ("utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"):
        try:
            t = raw.decode(enc)
            break
        except Exception:
            continue
    else:
        t = raw.decode("utf-8", errors="replace")
    t = t.replace("\x00", " ")
    if len(t) > max_chars:
        t = t[:max_chars] + "\n…(이하 잘림)…"
    return t


def _pdf_to_text(raw: bytes, max_chars: int) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return "(PDF 파싱 라이브러리를 사용할 수 없습니다.)"
    try:
        r = PdfReader(io.BytesIO(raw))
        parts: list[str] = []
        remain = max_chars
        for page in r.pages:
            if remain <= 0:
                break
            try:
                txt = (page.extract_text() or "").strip()
            except Exception:
                txt = ""
            if not txt:
                continue
            if len(txt) > remain:
                txt = txt[:remain] + "\n…(페이지 잘림)…"
            parts.append(txt)
            remain -= len(txt)
        return "\n\n".join(parts).strip() if parts else "(PDF에서 추출한 텍스트가 없습니다.)"
    except Exception as e:
        return f"(PDF 읽기 오류: {e})"


def _xlsx_outline(raw: bytes, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "(Excel 파싱 라이브러리를 사용할 수 없습니다.)"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        out: list[str] = []
        used = 0
        for sheet_name in wb.sheetnames:
            if used >= max_chars:
                break
            ws = wb[sheet_name]
            hdr = [
                "--- 시트 이름 ---",
                sheet_name,
                "--- 각 열 헤더/샘플(최대 15행) ---",
            ]
            lines: list[str] = list(hdr)
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, max_col=40, values_only=True), start=1):
                cells = []
                for c in row:
                    if c is None:
                        cells.append("")
                    else:
                        cells.append(str(c).replace("\n", " ").strip()[:120])
                lines.append(f"행{ri}: " + " | ".join(cells))
            block = "\n".join(lines)
            if used + len(block) > max_chars:
                block = block[: max_chars - used] + "\n…"
            out.append(block)
            used += len(block)
            if used >= max_chars:
                break
        wb.close()
        return "\n\n".join(out) if out else "(시트 정보를 읽지 못했습니다.)"
    except Exception as e:
        return f"(Excel 읽기 오류: {e})"


def _one_file_digest(filename: str, raw: bytes, per_budget: int) -> str:
    fn = (filename or "file").strip()
    ext = os.path.splitext(fn)[1].lower()
    if ext in _TEXT_EXT:
        return f"[{fn}]\n{_decode_text(raw, per_budget)}"
    if ext in _EXCEL_EXT:
        return f"[{fn} — Excel 서버 추출 요약]\n{_xlsx_outline(raw, per_budget)}"
    if ext in _PDF_EXT:
        return f"[{fn} — PDF 텍스트 추출]\n{_pdf_to_text(raw, per_budget)}"
    return f"[{fn}] 자동 추출 대상 형식 아님(텍스트·Excel·PDF만 서버 추출)."


def build_attachment_llm_digest(
    entries: list[dict[str, Any]],
    *,
    max_total_chars: int = 12_000,
    note: str = "",
) -> str:
    """
    저장된 첨부 경로에서 바이너리를 읽어 LLM 프롬프트 문자열 생성.
    """
    if not entries:
        return ""
    budgets = max_total_chars // max(1, len(entries))
    budgets = max(budgets, 1_800)
    parts: list[str] = []
    if note.strip():
        parts.append(note.strip())
    parts.append("(아래는 서버가 저장한 첨부에서 추출했습니다.)")
    for ent in entries:
        path = (ent or {}).get("path") or ""
        fname = ((ent or {}).get("filename") or "file").strip()
        memo = ((ent or {}).get("note") or "").strip()
        raw = r2_storage.read_bytes_from_ref(path)
        if not raw:
            parts.append(f"[{fname}] — 파일 바이너리를 읽지 못했습니다.")
            continue
        hdr = fname
        if memo:
            hdr += f" (메모: {memo})"
        parts.append(hdr)
        parts.append(_one_file_digest(fname, raw, budgets))
    text = "\n\n".join(parts)
    if len(text) > max_total_chars:
        text = text[:max_total_chars] + "\n…(첨부 컨텍스트 상한)…"
    return text.strip()
